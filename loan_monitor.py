"""
============================================================
  LOAN MONITORING SYSTEM
  Extracts loans due within 7 days from SQL database,
  generates a formatted Excel report, and emails it to
  the Credit & Loans Department via SMTP.
============================================================
"""

import sqlite3
import smtplib
import os
import configparser
import logging
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Logging setup ────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler("loan_monitor.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
#  1. CONFIGURATION
# ─────────────────────────────────────────────────────────────

def load_config(path="config.ini") -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Config file not found: {path}")
    cfg.read(path)
    return cfg


# ─────────────────────────────────────────────────────────────
#  2. DATABASE — SETUP & QUERY
# ─────────────────────────────────────────────────────────────

def setup_demo_database(db_path: str):
    """Creates and seeds a demo SQLite database from the SQL schema file."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    with open("loan_schema.sql", "r") as f:
        sql = f.read()
    # Execute statement by statement
    for statement in sql.split(";"):
        stmt = statement.strip()
        if stmt:
            try:
                cursor.execute(stmt)
            except sqlite3.OperationalError:
                pass  # Table/data already exists
    conn.commit()
    conn.close()
    log.info(f"Demo database ready: {db_path}")


def get_db_connection(cfg: configparser.ConfigParser):
    db_type = cfg.get("DATABASE", "db_type", fallback="sqlite").lower()

    if db_type == "sqlite":
        db_path = cfg.get("DATABASE", "db_path", fallback="loans.db")
        if not os.path.exists(db_path):
            log.info("Database not found – creating demo database...")
            setup_demo_database(db_path)
        return sqlite3.connect(db_path)

    elif db_type == "mysql":
        import mysql.connector
        return mysql.connector.connect(
            host=cfg.get("DATABASE", "db_host"),
            port=cfg.getint("DATABASE", "db_port", fallback=3306),
            database=cfg.get("DATABASE", "db_name"),
            user=cfg.get("DATABASE", "db_user"),
            password=cfg.get("DATABASE", "db_password")
        )

    elif db_type == "postgresql":
        import psycopg2
        return psycopg2.connect(
            host=cfg.get("DATABASE", "db_host"),
            port=cfg.getint("DATABASE", "db_port", fallback=5432),
            dbname=cfg.get("DATABASE", "db_name"),
            user=cfg.get("DATABASE", "db_user"),
            password=cfg.get("DATABASE", "db_password")
        )

    else:
        raise ValueError(f"Unsupported db_type: {db_type}")


def fetch_due_loans(conn, days_ahead: int = 7, include_overdue: bool = True) -> pd.DataFrame:
    """
    Pulls loans due within `days_ahead` days (and optionally overdue loans).
    Returns a DataFrame with all required report fields.
    """
    today = date.today()
    cutoff = today + timedelta(days=days_ahead)

    if include_overdue:
        due_filter = "l.due_date <= :cutoff"
    else:
        due_filter = "l.due_date BETWEEN :today AND :cutoff"

    query = f"""
        SELECT
            c.full_name            AS "Customer Name",
            l.loan_id              AS "Loan ID",
            l.amount_borrowed      AS "Amount Borrowed",
            l.outstanding_balance  AS "Outstanding Balance",
            l.due_date             AS "Due Date",
            c.phone_number         AS "Phone Number",
            c.email                AS "Email",
            b.loan_officer         AS "Loan Officer",
            b.branch_name          AS "Branch",
            l.loan_status          AS "Loan Status"
        FROM loans l
        JOIN customers c ON l.customer_id = c.customer_id
        JOIN branches  b ON l.branch_id   = b.branch_id
        WHERE {due_filter}
          AND l.loan_status != 'Paid'
        ORDER BY l.due_date ASC
    """

    params = {"cutoff": str(cutoff), "today": str(today)}
    df = pd.read_sql_query(query, conn, params=params)

    # Calculate Days Remaining (negative = overdue)
    df["Due Date"] = pd.to_datetime(df["Due Date"]).dt.date
    df["Days Remaining"] = df["Due Date"].apply(lambda d: (d - today).days)

    # Combine Loan Officer + Branch
    df["Loan Officer / Branch"] = df["Loan Officer"] + " / " + df["Branch"]
    df.drop(columns=["Loan Officer", "Branch"], inplace=True)

    # Reorder columns to match spec
    df = df[[
        "Customer Name",
        "Loan ID",
        "Amount Borrowed",
        "Outstanding Balance",
        "Due Date",
        "Days Remaining",
        "Phone Number",
        "Email",
        "Loan Officer / Branch",
        "Loan Status"
    ]]

    log.info(f"Fetched {len(df)} loans due within {days_ahead} days.")
    return df


# ─────────────────────────────────────────────────────────────
#  3. EXCEL REPORT GENERATION
# ─────────────────────────────────────────────────────────────

# Colour palette
CLR_HEADER_BG    = "1F3864"   # Dark navy
CLR_HEADER_FG    = "FFFFFF"   # White
CLR_OVERDUE_BG   = "FFE0E0"   # Light red
CLR_ACTIVE_BG    = "E8F5E9"   # Light green
CLR_ALT_ROW      = "F5F8FF"   # Very light blue
CLR_TITLE_BG     = "2E75B6"   # Medium blue
CLR_SUMMARY_BG   = "EBF3FB"   # Pale blue

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")


def generate_excel_report(df: pd.DataFrame, company_name: str) -> str:
    """Builds a styled Excel report and returns the file path."""
    today = date.today()
    report_dir = "reports"
    os.makedirs(report_dir, exist_ok=True)
    filename = f"{report_dir}/Loans_Due_Report_{today.strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Loan Detail ──────────────────────────────────
    ws = wb.active
    ws.title = "Loans Due This Week"
    ws.sheet_view.showGridLines = False

    col_count = len(df.columns)

    # --- Title Block ---
    ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
    title_cell = ws["A1"]
    title_cell.value = f"📋  {company_name.upper()}  –  LOAN DUE DATE ALERT"
    title_cell.font = _font(bold=True, color=CLR_HEADER_FG, size=14)
    title_cell.fill = _fill(CLR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(col_count)}2")
    sub_cell = ws["A2"]
    sub_cell.value = (
        f"Report Generated: {today.strftime('%A, %d %B %Y')}  |  "
        f"Loans Due: {today.strftime('%d %b')} – {(today + timedelta(days=7)).strftime('%d %b %Y')}  |  "
        f"Total Records: {len(df)}"
    )
    sub_cell.font = _font(italic=True, color="FFFFFF", size=9)
    sub_cell.fill = _fill("2E75B6")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Blank spacer row
    ws.row_dimensions[3].height = 6

    # --- Column Headers (row 4) ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = _font(bold=True, color=CLR_HEADER_FG, size=10)
        cell.fill = _fill(CLR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 32

    # --- Data Rows ---
    for row_idx, row in enumerate(df.itertuples(index=False), start=5):
        status = row._9  # "Loan Status" column (0-indexed position 9)
        days_remaining = row._5  # "Days Remaining" column

        # Row background: overdue = red tint, active = alternating
        if status == "Overdue" or days_remaining < 0:
            row_fill = _fill(CLR_OVERDUE_BG)
        elif row_idx % 2 == 0:
            row_fill = _fill(CLR_ALT_ROW)
        else:
            row_fill = _fill("FFFFFF")

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            col_name = df.columns[col_idx - 1]

            if col_name in ("Amount Borrowed", "Outstanding Balance"):
                cell.value = value
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "Due Date":
                cell.value = str(value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "Days Remaining":
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if value < 0:
                    cell.font = _font(bold=True, color="CC0000", size=10)
                elif value <= 3:
                    cell.font = _font(bold=True, color="E65100", size=10)
                else:
                    cell.font = _font(size=10)
            elif col_name == "Loan Status":
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if value == "Overdue":
                    cell.font = _font(bold=True, color="CC0000", size=10)
                else:
                    cell.font = _font(bold=True, color="1B7A1B", size=10)
            else:
                cell.value = value
                cell.font = _font(size=10)

        ws.row_dimensions[row_idx].height = 20

    # --- Column Widths ---
    col_widths = {
        "Customer Name": 22,
        "Loan ID": 15,
        "Amount Borrowed": 18,
        "Outstanding Balance": 20,
        "Due Date": 13,
        "Days Remaining": 15,
        "Phone Number": 18,
        "Email": 28,
        "Loan Officer / Branch": 28,
        "Loan Status": 14,
    }
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

    # Freeze header rows
    ws.freeze_panes = "A5"

    # ── Sheet 2: Summary ─────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    summary_title = ws2["A1"]
    summary_title.value = "LOAN REPORT SUMMARY"
    summary_title.font = _font(bold=True, color=CLR_HEADER_FG, size=13)
    summary_title.fill = _fill(CLR_TITLE_BG)
    summary_title.alignment = Alignment(horizontal="center", vertical="center")
    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 28

    overdue_df = df[df["Loan Status"] == "Overdue"]
    active_df  = df[df["Loan Status"] == "Active"]

    summary_data = [
        ("Report Date",                today.strftime("%d %B %Y")),
        ("Reporting Period",            f"Next 7 Days + Overdue"),
        ("", ""),
        ("LOAN COUNTS", ""),
        ("Total Loans in Report",       len(df)),
        ("Active Loans",                len(active_df)),
        ("Overdue Loans",               len(overdue_df)),
        ("", ""),
        ("FINANCIAL SUMMARY", ""),
        ("Total Amount Borrowed (KES)", f"{df['Amount Borrowed'].sum():,.2f}"),
        ("Total Outstanding (KES)",     f"{df['Outstanding Balance'].sum():,.2f}"),
        ("Overdue Outstanding (KES)",   f"{overdue_df['Outstanding Balance'].sum():,.2f}"),
        ("", ""),
        ("DUE THIS WEEK", ""),
        ("Due Today",                   len(df[df["Days Remaining"] == 0])),
        ("Due in 1-3 Days",             len(df[(df["Days Remaining"] >= 1) & (df["Days Remaining"] <= 3)])),
        ("Due in 4-7 Days",             len(df[(df["Days Remaining"] >= 4) & (df["Days Remaining"] <= 7)])),
        ("Already Overdue",             len(df[df["Days Remaining"] < 0])),
    ]

    for r_idx, (label, value) in enumerate(summary_data, start=2):
        label_cell = ws2.cell(row=r_idx, column=1, value=label)
        value_cell = ws2.cell(row=r_idx, column=2, value=value)

        if label in ("LOAN COUNTS", "FINANCIAL SUMMARY", "DUE THIS WEEK"):
            label_cell.font = _font(bold=True, color=CLR_HEADER_FG, size=10)
            value_cell.font = _font(bold=True, color=CLR_HEADER_FG, size=10)
            label_cell.fill = _fill(CLR_HEADER_BG)
            value_cell.fill = _fill(CLR_HEADER_BG)
        elif label == "":
            label_cell.fill = _fill("FFFFFF")
            value_cell.fill = _fill("FFFFFF")
        else:
            bg = CLR_SUMMARY_BG if r_idx % 2 == 0 else "FFFFFF"
            label_cell.fill = _fill(bg)
            value_cell.fill = _fill(bg)
            label_cell.font = _font(size=10)
            value_cell.font = _font(bold=True, size=10)

        label_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r_idx].height = 20

    wb.save(filename)
    log.info(f"Excel report saved: {filename}")
    return filename


# ─────────────────────────────────────────────────────────────
#  4. EMAIL — BUILD & SEND
# ─────────────────────────────────────────────────────────────

def build_email_body(df: pd.DataFrame, company_name: str) -> str:
    today = date.today()
    overdue_count  = len(df[df["Loan Status"] == "Overdue"])
    active_count   = len(df[df["Loan Status"] == "Active"])
    total_outstanding = df["Outstanding Balance"].sum()
    overdue_outstanding = df[df["Loan Status"] == "Overdue"]["Outstanding Balance"].sum()

    rows_html = ""
    for _, row in df.iterrows():
        color = "#ffe0e0" if row["Loan Status"] == "Overdue" else ("#ffffff" if _ % 2 == 0 else "#f5f8ff")
        days_display = (
            f"<span style='color:red;font-weight:bold;'>{row['Days Remaining']} (Overdue)</span>"
            if row["Days Remaining"] < 0
            else f"<span style='color:{'#e65100' if row['Days Remaining'] <= 3 else '#1b7a1b'};font-weight:bold;'>{row['Days Remaining']}</span>"
        )
        status_color = "red" if row["Loan Status"] == "Overdue" else "green"
        rows_html += f"""
        <tr style="background-color:{color};">
            <td>{row['Customer Name']}</td>
            <td>{row['Loan ID']}</td>
            <td style='text-align:right;'>KES {row['Amount Borrowed']:,.2f}</td>
            <td style='text-align:right;'>KES {row['Outstanding Balance']:,.2f}</td>
            <td style='text-align:center;'>{row['Due Date']}</td>
            <td style='text-align:center;'>{days_display}</td>
            <td>{row['Phone Number']}</td>
            <td>{row['Loan Officer / Branch']}</td>
            <td style='color:{status_color};font-weight:bold;text-align:center;'>{row['Loan Status']}</td>
        </tr>"""

    return f"""
    <html><body style="font-family:Arial, sans-serif; color:#222; font-size:14px;">
      <div style="background:#1F3864;color:white;padding:20px 30px;border-radius:6px 6px 0 0;">
        <h2 style="margin:0;">📋 {company_name}</h2>
        <p style="margin:5px 0 0;">Loan Due Date Alert — Loans Due Within 7 Days</p>
        <p style="margin:4px 0 0;font-size:12px;opacity:0.8;">Generated: {today.strftime('%A, %d %B %Y')}</p>
      </div>

      <div style="background:#EBF3FB;padding:16px 30px;border:1px solid #cce0f5;">
        <h3 style="margin:0 0 10px;color:#1F3864;">Report Summary</h3>
        <table style="border-collapse:collapse;width:100%;max-width:600px;">
          <tr>
            <td style="padding:6px 12px;background:#fff;border:1px solid #ddd;"><b>Total Loans in Report</b></td>
            <td style="padding:6px 12px;background:#fff;border:1px solid #ddd;text-align:center;"><b>{len(df)}</b></td>
          </tr>
          <tr>
            <td style="padding:6px 12px;background:#e8f5e9;border:1px solid #ddd;">Active Loans</td>
            <td style="padding:6px 12px;background:#e8f5e9;border:1px solid #ddd;text-align:center;color:green;font-weight:bold;">{active_count}</td>
          </tr>
          <tr>
            <td style="padding:6px 12px;background:#ffe0e0;border:1px solid #ddd;">Overdue Loans</td>
            <td style="padding:6px 12px;background:#ffe0e0;border:1px solid #ddd;text-align:center;color:red;font-weight:bold;">{overdue_count}</td>
          </tr>
          <tr>
            <td style="padding:6px 12px;background:#fff;border:1px solid #ddd;">Total Outstanding Balance</td>
            <td style="padding:6px 12px;background:#fff;border:1px solid #ddd;text-align:center;"><b>KES {total_outstanding:,.2f}</b></td>
          </tr>
          <tr>
            <td style="padding:6px 12px;background:#ffe0e0;border:1px solid #ddd;">Overdue Outstanding</td>
            <td style="padding:6px 12px;background:#ffe0e0;border:1px solid #ddd;text-align:center;color:red;font-weight:bold;">KES {overdue_outstanding:,.2f}</td>
          </tr>
        </table>
      </div>

      <div style="padding:20px 30px;">
        <h3 style="color:#1F3864;">Loan Details</h3>
        <table style="border-collapse:collapse;width:100%;font-size:12px;">
          <thead>
            <tr style="background:#1F3864;color:white;">
              <th style="padding:8px;text-align:left;">Customer Name</th>
              <th style="padding:8px;text-align:left;">Loan ID</th>
              <th style="padding:8px;text-align:right;">Amount Borrowed</th>
              <th style="padding:8px;text-align:right;">Outstanding Balance</th>
              <th style="padding:8px;text-align:center;">Due Date</th>
              <th style="padding:8px;text-align:center;">Days Remaining</th>
              <th style="padding:8px;text-align:left;">Phone</th>
              <th style="padding:8px;text-align:left;">Loan Officer / Branch</th>
              <th style="padding:8px;text-align:center;">Status</th>
            </tr>
          </thead>
          <tbody>
            {rows_html}
          </tbody>
        </table>
      </div>

      <div style="background:#f0f0f0;padding:12px 30px;border-top:2px solid #1F3864;font-size:11px;color:#666;">
        ⚠️ This is an automated report. Please do not reply to this email.
        Full details are in the attached Excel file. &nbsp;|&nbsp; {company_name} — Loan Monitoring System
      </div>
    </body></html>
    """


def send_email(cfg: configparser.ConfigParser, report_path: str, df: pd.DataFrame):
    sender_email    = cfg.get("EMAIL", "sender_email")
    sender_password = cfg.get("EMAIL", "sender_password")
    recipient_email = cfg.get("EMAIL", "recipient_email")
    recipient_name  = cfg.get("EMAIL", "recipient_name", fallback="Credit and Loans Department")
    smtp_server     = cfg.get("EMAIL", "smtp_server", fallback="smtp.gmail.com")
    smtp_port       = cfg.getint("EMAIL", "smtp_port", fallback=587)
    subject         = cfg.get("EMAIL", "email_subject", fallback="Loan Due Date Alert")
    company_name    = cfg.get("EMAIL", "company_name", fallback="Loan Management System")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"{subject} – {date.today().strftime('%d %b %Y')}"
    msg["From"]    = f"{company_name} <{sender_email}>"
    msg["To"]      = f"{recipient_name} <{recipient_email}>"

    # Attach HTML body
    html_body = build_email_body(df, company_name)
    msg.attach(MIMEText(html_body, "html"))

    # Attach Excel file
    with open(report_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename={os.path.basename(report_path)}"
    )
    msg.attach(part)

    # Send via SMTP
    log.info(f"Connecting to {smtp_server}:{smtp_port}...")
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.ehlo()
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_email, msg.as_string())

    log.info(f"✅  Email sent successfully to {recipient_email}")


# ─────────────────────────────────────────────────────────────
#  5. MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────

def main():
    log.info("=" * 55)
    log.info("  LOAN MONITORING SYSTEM — Starting Run")
    log.info("=" * 55)

    # Load config
    cfg = load_config("config.ini")
    days_ahead      = cfg.getint("REPORT", "days_lookahead", fallback=7)
    include_overdue = cfg.getboolean("REPORT", "include_overdue", fallback=True)
    save_local_copy = cfg.getboolean("REPORT", "save_local_copy", fallback=True)
    company_name    = cfg.get("EMAIL", "company_name", fallback="Loan Management System")

    # Connect to DB and fetch loans
    conn = get_db_connection(cfg)
    df = fetch_due_loans(conn, days_ahead=days_ahead, include_overdue=include_overdue)
    conn.close()

    if df.empty:
        log.info("No loans due within the reporting window. No email sent.")
        return

    # Generate Excel report
    report_path = generate_excel_report(df, company_name)

    # Send email
    try:
        send_email(cfg, report_path, df)
    except Exception as e:
        log.error(f"❌  Failed to send email: {e}")
        log.info(f"Report saved locally at: {report_path}")
        raise

    # Clean up local copy if not needed
    if not save_local_copy and os.path.exists(report_path):
        os.remove(report_path)
        log.info("Local report copy removed (save_local_copy=false).")

    log.info("=" * 55)
    log.info("  Run complete.")
    log.info("=" * 55)


if __name__ == "__main__":
    main()
